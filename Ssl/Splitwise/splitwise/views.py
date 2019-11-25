from django.shortcuts import render, redirect, HttpResponse

# Create your views here.
from ipywidgets import widgets
from datetime import datetime,date
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
import plotly.graph_objects as go
from django.shortcuts import render
from plotly.offline import plot
from plotly.subplots import make_subplots
import numpy as np
import pandas as pd
import scipy as sp


from django.contrib.auth.forms import UserCreationForm
from django.urls import reverse_lazy
from django.views import generic
from django.http import HttpResponse, HttpResponseRedirect
from splitwise.forms import CustomUserCreationForm
from splitwise.forms import *

from .models import *
from django.template import loader

from django.template import RequestContext
from django.db.models import Q

no_transactions = 0

class SignUp(generic.CreateView):
    form_class = CustomUserCreationForm
    success_url = reverse_lazy('login')
    template_name = 'signup.html'

class EditProfile(generic.CreateView):
	form_class = ProfileUpdateForm
	success_url = reverse_lazy('login')
	template_name = 'editprofile.html'

#def edit_profile_view(request): 
  
#	if request.method == 'POST': 
#		form = ProfileUpdateForm(request.POST, instance=request.user) 
  
#		if form.is_valid(): 
#			form.save() 
#			return redirect('success') 
#	else:
#		form = ProfileUpdateForm() 
#		return render(request, 'editprofile.html', {'form' : form}) 
  
  
def success(request):
	me = User.objects.get(username=request.user.get_username())
	usr = request.user.get_username()
	RANGE_SHORT = 's'
	RANGE_MID = 'm'
	RANGE_LONG = 'l'
	RANGE_CHOICES = (
		(RANGE_SHORT, 'Short '),
		(RANGE_MID, 'Mid rbjkiange'),
		(RANGE_LONG, 'Long range')
	)

	friend_form = FriendForm()
	
	xyz=Friend.objects.filter(person1=me)
	#print(xyz)
	final_choices=()
	for e in xyz:
		#print(str(e.person2))
		thistuple = (str(e.person2), str(e.person2))
		#print(thistuple)
		final_choices = final_choices + (thistuple,)
	#print(final_choices)
	group_form = GroupForm(final_choices)
	#group_form.fields['friends'].choices=RANGE_CHOICES
	if request.method == 'POST':
		if 'friend' in request.POST:
			friend_form = FriendForm(request.POST)
			if friend_form.is_valid():
				friend_id = friend_form.cleaned_data['your_name']
				if User.objects.filter(username=friend_id).exists():
					friend = User.objects.get(username=friend_id)
					if friend == me:
						print('F')
					elif Friend.objects.filter(person1=me,person2=friend).exists() or Friend.objects.filter(person1=friend,person2=me).exists():
						print("f")
					else:
						f = Friend(person1=me,person2=friend)
						f1 = Friend(person1=friend,person2=me)

						print('hi')
						f.save()
						f1.save()
				return HttpResponseRedirect("/splitwise/success/")

						
		if 'group' in request.POST:
			
			group_form = GroupForm(final_choices, request.POST)
			if group_form.is_valid():
				name = group_form.cleaned_data['group_name']
				people = group_form.cleaned_data['friends']
				g = Group(group_name=name)
				g.save()
				m = Membership(friend = me, group =g)
				m.save()
				for p in people:
					preal = User.objects.get(username=p)
					for p1 in people:
						if p!=p1:
							p1real = User.objects.get(username=p1)
							if not Friend.objects.filter(person1=preal,person2=p1real).exists():
								fxxx = Friend(person1=preal,person2=p1real)
								fxxx1 = Friend(person1=p1real,person2=preal)
								fxxx.save()
								fxxx1.save()
				for p in people:
					member = User.objects.get(username=p)
					m = Membership(friend = member, group = g)
					m.save()
					#print(member)
				return HttpResponseRedirect("/splitwise/success/") 
	else:
		friend_form=FriendForm()
		group_form=GroupForm(final_choices)
		#group_form.fields['friends'].choices=[]
	
	friends = Friend.objects.filter(person1=me)
	friends_boolean=[]
	for y in friends:
		if(y.money_owed<0):
			y.money_owed = (-1)*(y.money_owed)
			friends_boolean.append(0)
		else:
			friends_boolean.append(1)
	friends_list = zip(friends,friends_boolean)
	groups = Membership.objects.filter(friend=me)
	groups_boolean=[]
	for g in groups:
		if(g.money_owed<0):
			g.money_owed=-1*g.money_owed
			groups_boolean.append(0)
		else:
			groups_boolean.append(1)
	print(groups)
	groups_list = zip(groups,groups_boolean)
	edit_profile_form = ProfileUpdateForm()
	if request.method == 'POST':
		if 'edit_profile' in request.POST:
			print(request.user.profile.bio)
			edit_profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
			if edit_profile_form.is_valid():
				#edit_profile_form.save()
				#bio = edit_profile_form.cleaned_data['bio']
				#print(bio)
				#image = edit_profile_form.cleaned_data['image']
				#print(request.FILES['image'])
				#z = Profile.objects.get(user = me)
				#z.bio = bio
				#z.image = "profile_image/"+str(request.FILES['image'])
				#print(Profile.objects.get(user = me).image)
				#z.save()
				edit_profile_form.save()
				#print(bio)
	else:
		edit_profile_form=ProfileUpdateForm()
	
	context = {
		'friend_form' : friend_form,
		'group_form' : group_form,
		'friends_list' : friends_list,
		'groups_list' : groups_list,
		'edit_profile_form' : edit_profile_form
	}
	

	template = loader.get_template('home.html') 
	#return render(request, 'home.html',
         #  context_instance=RequestContext(request))
	return HttpResponse(template.render(context, request))

def friend(request,f):
	me = request.user
	print(f)
	friend = User.objects.get(id=f)
	print(friend)
	x = Friend.objects.filter(person1__username=me)
	print(f)
	z=0
	a=''
	for y in x:
		if(y.person2.username==friend.username):
			zxxx = 0
			ts = Transaction.objects.filter(lender=me,borrower=friend,group=None)
			for t in ts:
				zxxx = zxxx + t.amount
			ts = Transaction.objects.filter(lender=friend,borrower=me,group=None)
			for t in ts:
				zxxx = zxxx - t.amount
			a = str(y.person2)
	groups_list = []
	groups = Membership.objects.filter(friend=me)
	for g in groups:
		if Membership.objects.filter(group=g.group, friend=friend).exists():
			m = Membership.objects.get(group=g.group, friend=friend)
			groups_list.append([m.group,0])

	for g in groups_list:
		tlist = Transaction.objects.filter(group=g[0],lender=me,borrower=friend)
		for t in tlist:
			g[1]=g[1]+t.amount
		tlist = Transaction.objects.filter(group=g[0],lender=friend,borrower=me)
		for t in tlist:
			g[1]=g[1]-t.amount
	print(groups_list)
	for g in groups_list:
		print(g[0].group_name)
	if request.method == 'POST':
		if 'settle_up' in request.POST:
			if zxxx>0:
				t = Transaction(group_transaction_id=Transaction.no_transactions,lender=friend,borrower=me,description='Settling!',amount=zxxx,tag='st',added_by=me,paid_by=friend)
				t.save()
				Transaction.no_transactions = Transaction.no_transactions + 1
			elif zxxx<0:
				t = Transaction(group_transaction_id=Transaction.no_transactions,lender=me,borrower=friend,description='Settling!',amount=-1*zxxx,tag='st',added_by=me,paid_by=friend)
				t.save()
				Transaction.no_transactions = Transaction.no_transactions + 1
			else:
				pass
			for g in groups_list:
				if g[1] > 0:
					no = g[0].no_transactions
					t = Transaction(group=g[0],group_transaction_id=no,lender=friend,borrower=me,description='Settling!',amount=g[1],tag='st',added_by=me,paid_by=friend)
					t.save()
					m1 = Membership.objects.get(group=g[0],friend=me)
					z = m1.money_owed - g[1]
					m1.money_owed = z
					m1.save()
					m2 = Membership.objects.get(group=g[0],friend=friend)
					z = m2.money_owed + g[1]
					m2.money_owed = z
					m2.save()
					z = g[0].no_transactions + 1
					g[0].no_transactions = z
					g[0].save()
				elif g[1] < 0:
					no = g[0].no_transactions
					t = Transaction(group=g[0],group_transaction_id=no,lender=me,borrower=friend,description='Settling!',amount=-1*g[1],tag='st',added_by=me,paid_by=friend)
					t.save()
					m1 = Membership.objects.get(group=g[0],friend=friend)
					z = m1.money_owed - (-1*g[1])
					m1.money_owed = z
					m1.save()
					m2 = Membership.objects.get(group=g[0],friend=me)
					z = m2.money_owed + (-1*g[1])
					m2.money_owed = z
					m2.save()
					z = g[0].no_transactions + 1
					g[0].no_transactions = z
					g[0].save()
					pass
				else:
					pass
		f1 = Friend.objects.get(person1=me,person2=friend)
		f1.money_owed = 0
		f1.save()
		f2 = Friend.objects.get(person1=friend,person2=me)
		f2.money_owed = 0
		f2.save()
		idx = str(friend.id)
		return HttpResponseRedirect('/splitwise/success/')
	template = loader.get_template('expanded_friend.html')
	print(zxxx)
	if(zxxx >= 0):
		boolean=1
	else:
		boolean=0
		zxxx = (-1)*zxxx
	boolean2=[]
	for g in groups_list:
		if(g[1] >= 0):
			boolean2.append(1)
		else:
			boolean2.append(0)
			g[1] = (-1)*g[1]
	lst = zip(groups_list,boolean2)
	context = {
		'zxxx' : zxxx,
		'boolean' : boolean,
		'a' : a,
		'friend':friend,
		'lst':lst
		#'groups_list':groups_list
	}
	return HttpResponse(template.render(context, request))

def tag(argument): 
    switcher = { 
        0: "zero", 
        1: "one", 
        2: "two", 
        'mv': 'Movies',
		'fd': 'Food',
		'tr': 'Travel',
		'ee': 'Electronics',
		'md': 'Medical',
		'sp': 'Shopping',
		'sv': 'Services',
		'st': 'Settle',
		'ot': 'Others'
    } 

  
    # get() method of dictionary data type returns  
    # value of passed argument if it is present  
    # in dictionary otherwise second argument will 
    # be assigned as default value of passed argument 
    return switcher.get(argument, "nothing") 

def group(request,g):
	me=request.user
	group = Group.objects.get(id=g)
	all_transactions = Transaction.objects.filter(group=g).order_by('date')
	no =group.no_transactions
	ms = Membership.objects.filter(group=g)
	frnd_list=()
	for m in ms:
		if m.friend != me:
			frnd_list = frnd_list + ((m.friend.username, m.friend.username),)
	template = loader.get_template('expanded_group.html')
	x = ''
	trans_list=[]
	request.session['group']=g
	for i in range(no):
		if all_transactions.filter(group_transaction_id=i).exists():
			ts=all_transactions.filter(group_transaction_id=i)
			shares_list = []
			paid_amt=0
			for t in ts:
				date=[str(t.date)]
				paid_by=[t.paid_by.username]
				x = tag(t.tag)
				tag1=[x]
				desc=[t.description]
				contri=t.amount
				paid_amt=paid_amt+contri
				borrower=t.borrower.username
				shares_list.append([borrower,contri])
			lst = (date,paid_by,tag1,desc,shares_list,paid_amt)
			trans_list.append(lst)

	change_form = ChangeForm(frnd_list)
	if request.method == 'POST':
		if 'balances' in request.POST:
			print('hi')
		if 'settle_up' in request.POST:
			change_form = ChangeForm(frnd_list,request.POST)
			if change_form.is_valid():
				people = change_form.cleaned_data['friends']
				for p in people:
					f = User.objects.get(username=p)
					amt=0
					if all_transactions.filter(lender=me,borrower=f).exists():
						ts=all_transactions.filter(lender=me,borrower=f)
						for t in ts:
							amt = amt + t.amount
					if all_transactions.filter(lender=f,borrower=me).exists():
						ts=all_transactions.filter(lender=f,borrower=me)
						for t in ts:
							amt = amt - t.amount
					if amt>0:
						no = group.no_transactions
						t = Transaction(group=group,group_transaction_id=no,lender=f,borrower=me,description='Settling!',amount=amt,tag='st',added_by=me,paid_by=friend)
						t.save()
						m1 = Membership.objects.get(group=group,friend=me)
						z = m1.money_owed - amt
						m1.money_owed = z
						m1.save()
						m2 = Membership.objects.get(group=group,friend=f)
						z = m2.money_owed + amt
						m2.money_owed = z
						m2.save()
						f2 = Friend.objects.get(person1=me,person2=f)
						z = f2.money_owed - amt
						f2.money_owed=z
						f2.save()
						f1 = Friend.objects.get(person1=f,person2=me)
						z = f1.money_owed + amt
						f1.money_owed = z
						f1.save()
						z = group.no_transactions + 1
						group.no_transactions = z
						group.save()
					elif amt<0:
						no = group.no_transactions
						t = Transaction(group=group,group_transaction_id=no,lender=me,borrower=f,description='Settling!',amount=-1*amt,tag='st',added_by=me,paid_by=me)
						t.save()
						m1 = Membership.objects.get(group=group,friend=me)
						z = m1.money_owed + (-1*amt)
						m1.money_owed = z
						m1.save()
						m2 = Membership.objects.get(group=group,friend=f)
						z = m2.money_owed - (-1*amt)
						m2.money_owed = z
						m2.save()
						f2 = Friend.objects.get(person1=me,person2=f)
						z = f2.money_owed + (-1*amt)
						f2.money_owed = z
						f2.save()
						f1 = Friend.objects.get(person1=f,person2=me)
						z = f1.money_owed - (-1*amt)
						f1.money_owed = z
						f1.save()
						z = group.no_transactions + 1
						group.no_transactions = z
						group.save()
					else:
						pass
				return HttpResponseRedirect('/splitwise/success/')

	else: 
		change_form=ChangeForm(frnd_list)
	context = {
		'trans_list':trans_list,
		'g':g,
		'change_form':change_form
	}
	
	return HttpResponse(template.render(context, request))

def activity_tab(request):
	me=request.user	
	template = loader.get_template('activity_tab.html')
	x = Transaction.objects.filter(Q(lender=me) | Q(borrower=me)).order_by('-date')
	a = 0
	print(len(x))
	i=0
	desc=[]
	money=[]
	additional_info=[]
	transactions=[]
	boolean=[]
	no_of_activities=0
	exit=False
	while(i < len(x)):
		#print(x[i])
		if(x[i].lender==me and x[i].borrower==me):
			i = i+1
			pass
		else:
			if(x[i].lender==me):
				money.append(x[i].amount)
			else:
				money.append((-1)*x[i].amount)

			desc.append(x[i].description)
			transactions.append(x[i])
			if(x[i].group_id is None):
				additional_info.append("group does not exist xxxxxxwkebjkerjbkjengksdbkejbg")
			else:
				additional_info.append(x[i].group.group_name)
			j=i+1
			if(j>=len(x)):
				break
				#print(j)
			while(x[j].group_id == x[i].group_id and x[j].group_transaction_id == x[i].group_transaction_id):
				if( j < len(x)):
					if(x[j].lender==me and x[j].borrower==me):
						j = j+1
						pass
					else:
						if(x[j].lender==me):
							money[no_of_activities]+=x[j].amount
						else:
							money[no_of_activities]-=x[j].amount
						j=j+1
				else:
					exit=True
					break
			i=j
			no_of_activities = no_of_activities + 1
			if(exit == True):
				break
	#for i in range(len(desc)):
		#print(str(desc[i])+" "+str(money[i]))
		#print(x.transaction
	#send_list = [desc, money]#, additional_info]
	for i in range(money):
		if money[i] >= 0:
			boolean.append(1)
		else:
			boolean.append(1)
			money[i]=-1*money[i]
	send_list = zip(money,desc,additional_info,transactions,boolean)
	#print(send_list)
	context = {
		'send_list' : send_list
	}		
	return HttpResponse(template.render(context, request))

def detailed_activity1(request, i):
	y = Transaction.objects.filter(group=None, group_transaction_id=i)
	#print(y)
	person = []
	share = []
	payer = ''
	total_amount=0;
	if(len(y) == 1):
		payer = y[0].lender.username
		total_amount = y[0].amount
		person.append(y[0].borrower.username)
		share.append(total_amount)
	else:
		for z in y:
			person.append(z.borrower.username)
			share.append(z.amount)
			total_amount = total_amount + z.amount
			if(z.lender.username == z.borrower.username):
				payer = z.lender.username
	lst = zip(person,share)
	context = {
		'lst' : lst,
		'payer' : payer,
		'total_amount' : total_amount
	}
	template = loader.get_template('expanded_activity.html')
	return HttpResponse(template.render(context, request))
	#return HttpResponse('success')

def detailed_activity2(request, i, j):
	x = Group.objects.filter(id=i)
	y = Transaction.objects.filter(group=x[0], group_transaction_id=j)
	#print(y)
	person = []
	share = []
	payer = ''
	total_amount=0;
	for z in y:
		person.append(z.borrower.username)
		share.append(z.amount)
		total_amount = total_amount + z.amount
		if(z.lender.username == z.borrower.username):
			payer = z.lender.username
	lst = zip(person,share)
	context = {
		'lst' : lst,
		'payer' : payer,
		'total_amount' : total_amount
	}
	template = loader.get_template('expanded_activity.html')
	return HttpResponse(template.render(context, request))


def transaction(request):
	me = request.user
	xyz=Friend.objects.filter(person1=me)
	#print(xyz)
	final_choices=()
	for e in xyz:
		#print(str(e.person2))
		thistuple = (str(e.person2), str(e.person2))
		#print(thistuple)
		final_choices = final_choices + (thistuple,)
	involved_form = InvolvedForm(final_choices)
	if request.method=='POST':
		if 'involved' in request.POST:
			involved_form=InvolvedForm(final_choices, request.POST)
			if involved_form.is_valid():
				people = involved_form.cleaned_data['friends']
				for p in people:
					print(p)
				request.session['people'] = people
				request.session['choices'] = final_choices
				return HttpResponseRedirect('/splitwise/transaction/form/')
	else:
		involved_form=InvolvedForm(final_choices)
	template = loader.get_template('transaction.html')
	x=''
	context = {
		'involved_form':involved_form
	}
	return HttpResponse(template.render(context, request))

def transaction_form(request):
	me = request.user
	people = request.session.get('people')
	final_choices = ()
	final_choices = final_choices + ((me.username,me.username),)
	for p in people:
		thistuple = (str(p),str(p))
		final_choices = final_choices + (thistuple,)
	choices = request.session.get('choices')
	template = loader.get_template('transaction_form.html')
	transaction_form = TransactionForm(final_choices)
	if request.method == 'POST':
		if 'transaction' in request.POST:
			transaction_form=TransactionForm(final_choices, request.POST)
			if transaction_form.is_valid():
				desc = transaction_form.cleaned_data['description']
				who_paid = transaction_form.cleaned_data['who_paid']
				print(who_paid)
				amt = transaction_form.cleaned_data['amount']
				split = transaction_form.cleaned_data['split']
				print(split)
				tag = transaction_form.cleaned_data['tag']
				shares = {}
				for i in final_choices:
					data = transaction_form.cleaned_data[i[0]+' (%)']
					shares[str(i[0])]=data
				payer = User.objects.get(username=who_paid[0])
				
				if split == 'equal':
					share_amt = amt/len(final_choices)
					print(share_amt)
					t1 = Transaction(group_transaction_id=Transaction.no_transactions,lender=payer,borrower=payer,description=desc,amount=share_amt,tag=tag,added_by=me,paid_by=payer)
					t1.save()
					for p in final_choices:

						print(p[0])
						user = User.objects.get(username=p[0])
						if user!=payer:
							t = Transaction(group_transaction_id=Transaction.no_transactions,lender=payer,borrower=user,description=desc,amount=share_amt,tag=tag,added_by=me,paid_by=payer)
							t.save()
							f1 = Friend.objects.get(person1=payer,person2=user)
							x = f1.money_owed + share_amt
							f1.money_owed=x
							f1.save()
							f2 = Friend.objects.get(person1=user,person2=payer)
							y = f2.money_owed - share_amt
							f2.money_owed=y
							f2.save()


				else:
					for p in final_choices:
						share_amt = shares[str(p[0])]/100*amt
						user = User.objects.get(username=p[0])
						t = Transaction(group_transaction_id=Transaction.no_transactions,lender=payer,borrower=user,description=desc,amount=share_amt,tag=tag,added_by=me,paid_by=payer)
						t.save()
						if user!=payer:
							f1 = Friend.objects.get(person1=payer,person2=user)
							x = f1.money_owed + share_amt
							f1.money_owed=x
							f1.save()
							f2 = Friend.objects.get(person1=user,person2=payer)
							y = f2.money_owed - share_amt
							f2.money_owed=y
							f2.save()

						print(share_amt)
				Transaction.no_transactions = Transaction.no_transactions + 1
				return HttpResponseRedirect('/splitwise/success/')
	else:
		transaction_form = TransactionForm(final_choices)

	context ={
		'transaction_form' : transaction_form
	}
	return HttpResponse(template.render(context,request))

def group_transaction(request):
	me=request.user
	template = loader.get_template('group_transaction.html')
	group_id=request.session.get('group')
	print(group_id)
	this_group=Group.objects.get(id=group_id)
	print(this_group)
	final_choices=()
	members = Membership.objects.filter(group=this_group)
	print(members)
	for m in members:
		final_choices = final_choices + ((m.friend.username, m.friend.username),)
		
	print(final_choices)
	group_transaction_form = GroupTransactionForm(final_choices)
	if request.method == 'POST':
		if 'group_transaction' in request.POST:
			group_transaction_form=GroupTransactionForm(final_choices, request.POST)
			if group_transaction_form.is_valid():
				desc = group_transaction_form.cleaned_data['description']
				who_paid = group_transaction_form.cleaned_data['who_paid']
				print(who_paid)
				amt = group_transaction_form.cleaned_data['amount']
				split = group_transaction_form.cleaned_data['split']
				print(split)
				tag = group_transaction_form.cleaned_data['tag']
				no = this_group.no_transactions
				shares = {}
				for i in final_choices:
					data = group_transaction_form.cleaned_data[i[0]+' (%)']
					shares[str(i[0])]=data
				payer = User.objects.get(username=who_paid[0])
				
				if split == 'equal':
					share_amt = amt/len(final_choices)
					print(share_amt)
					t1 = Transaction(group=this_group,group_transaction_id=no,lender=payer,borrower=payer,description=desc,amount=share_amt,tag=tag,added_by=me,paid_by=payer)
					t1.save()
					
					for p in final_choices:

						print(p[0])
						user = User.objects.get(username=p[0])
						if user!=payer:
							t = Transaction(group=this_group,group_transaction_id=no,lender=payer,borrower=user,description=desc,amount=share_amt,tag=tag,added_by=me,paid_by=payer)
							t.save()
							m1 = Membership.objects.get(group=this_group,friend=user)
							z = m1.money_owed - share_amt
							m1.money_owed = z
							m1.save()
							m2 = Membership.objects.get(group=this_group,friend=payer)
							z = m2.money_owed + share_amt
							m2.money_owed = z
							m2.save()
							f1 = Friend.objects.get(person1=payer,person2=user)
							x = f1.money_owed + share_amt
							f1.money_owed=x
							f1.save()
							f2 = Friend.objects.get(person1=user,person2=payer)
							y = f2.money_owed - share_amt
							f2.money_owed=y
							f2.save()


				else:
					for p in final_choices:
						share_amt = shares[str(p[0])]/100*amt
						user = User.objects.get(username=p[0])
						t = Transaction(group=this_group,group_transaction_id=no,lender=payer,borrower=user,description=desc,amount=share_amt,tag=tag,added_by=me,paid_by=payer)
						t.save()
						if user!=payer:
							f1 = Friend.objects.get(person1=payer,person2=user)
							m1 = Membership.objects.get(group=this_group,friend=user)
							z = m1.money_owed - share_amt
							m1.money_owed = z
							m1.save()
							m2 = Membership.objects.get(group=this_group,friend=payer)
							z = m2.money_owed + share_amt
							m2.money_owed = z
							m2.save()
							x = f1.money_owed + share_amt
							f1.money_owed=x
							f1.save()
							f2 = Friend.objects.get(person1=user,person2=payer)
							y = f2.money_owed - share_amt
							f2.money_owed=y
							f2.save()
						
				


				x = this_group.no_transactions + 1
				this_group.no_transactions = x 
				this_group.save()
				return HttpResponseRedirect('/splitwise/success/')
	else:
		group_transaction_form = GroupTransactionForm(final_choices)
	context = {
		'group_transaction_form':group_transaction_form
	}
	return HttpResponse(template.render(context,request))

def balances(request):
	me = request.user
	template = loader.get_template('balances.html')
	g=request.session.get('group')
	group = Group.objects.get(id=g)
	print(group)
	lst = Membership.objects.filter(group=group)
	money = []
	frnds_list = []
	for l in lst:
		if l.friend != me:
			money.append([l.friend.username, l.money_owed, l.friend.id])
			frnds_list.append(l.friend)
	print(money)
	
	money_friends = []
	for f in frnds_list:
		tlist = Transaction.objects.filter(group=group,lender=me,borrower=f)
		amt = 0
		for t in tlist:
			amt=amt+t.amount
		tlist = Transaction.objects.filter(group=g[0],lender=f,borrower=me)
		for t in tlist:
			amt=amt-t.amount
		money_friends.append([f.username, amt, f.id])
	print(money_friends)
			
	x=''
	context = {
		'money': money,
		'money_friends':money_friends
	}
	return(HttpResponse(template.render(context,request)))

def Insights(request):
	me = request.user
	name=me.username
	s="Transactions_"+name+".xlsx"
	#me.username
	#q1 = Transaction.objects.all()
	# print("g \n")
	#print(q1)
	start_date=date(2019, 11, 11)
	end_date=date(2019, 11, 29)
	print(me)
	query_set = Transaction.objects.filter(Q(lender=me)|Q(borrower=me),date__range=(start_date, end_date)) # Poll.objects.get(Q(question__startswith='Who'),Q(pub_date=date(2005, 5, 2)) | Q(pub_date=date(2005, 5, 6)))
	print(query_set)
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = "Transaction History"
	fontObj1 = Font(name='Times New Roman', bold=True,color=colors.RED)
	sheet['A1'].font = fontObj1
	sheet['B1'].font = fontObj1
	sheet['C1'].font = fontObj1
	sheet['D1'].font = fontObj1
	sheet['E1'].font = fontObj1
	sheet['F1'].font = fontObj1
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 20
	sheet.column_dimensions['C'].width = 20
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 30
	sheet.column_dimensions['F'].width = 20
	sheet.cell(row=1, column=1).value = "Lender"
	sheet.cell(row=1, column=2).value = "Borrower"
	#group id if any to be inserted
	sheet.cell(row=1, column=3).value = "Amount"
	sheet.cell(row=1, column=4).value = "Type of Expense"
	sheet.cell(row=1, column=5).value = "Date and Time of Transaction"
	dict0={'mv': "Movies",
	'fd': "Food",
	'tr': "Travel",
	'ee': "Electronics",
	'md': "Medical",
	'sp': "Shopping",
	'sv': "Services",
	'ot': "Others"}
	rowNum=2
	for record in query_set:
		if(record.group is None):
			#colNum=2
			#for colNum in range(1, 7):
			sheet.cell(row=rowNum, column=1).value = record.lender.username
			sheet.cell(row=rowNum, column=2).value = record.borrower.username
			sheet.cell(row=rowNum, column=3).value = str(record.amount)
			sheet.cell(row=rowNum, column=4).value = dict0[record.tag]
			sheet.cell(row=rowNum, column=5).value = str(record.date)
			#print(record.lender.username+" "+record.borrower.username+" "+str(record.amount)+" "+record.tag+" "+str(record.date))
			rowNum=rowNum+1

		else:#change to print group name
			sheet.cell(row=rowNum, column=1).value = record.lender.username
			sheet.cell(row=rowNum, column=2).value = record.borrower.username
			sheet.cell(row=rowNum, column=3).value = str(record.amount)
			sheet.cell(row=rowNum, column=4).value = dict0[record.tag]
			sheet.cell(row=rowNum, column=5).value = str(record.date)
			#print(record.lender.username+" "+record.borrower.username+" "+str(record.amount)+" "+record.tag+" "+str(record.date))
			rowNum=rowNum+1

	wb.save(s)
		# plot_div = plot([Scatter(x=x_data, y=y_data,
 #                        mode='lines', name='test',
 #                        opacity=0.8, marker_color='green')],
 #               output_type='div')
	# fig = make_subplots(rows=2, cols=2,subplot_titles=("Plot 1", "Plot 2", "Plot 3", "Plot 4"), )
	# fig.add_trace(go.Figure(data=go.Bar(y=[2, 3, 1])))
	# 	#fig.write_html('first_figure.html', auto_open=True)
	# 	#plot_div = plot(fig, output_type='div')
	# labels = ['Oxygen','Hydrogen','Carbon_Dioxide','Nitrogen']
	# values = [4500, 2500, 1053, 500]
	# fig.add_trace(go.Figure(data=[go.Pie(labels=labels, values=values)]))
	# fig.add_trace(go.Figure(data=[go.Pie(labels=labels, values=values)]))
	# fig.add_trace(go.Figure(data=[go.Pie(labels=labels, values=values)]))

	v=np.zeros((5,9))
	dict2={'mv': 1,
	'fd': 2,
	'tr': 3,
	'ee': 4,
	'md': 5,
	'sp': 6,
	'sv': 7,
	'ot': 8}
	inverse_dict = {v: k for k, v in dict2.items()}
	dictionary= {"mv": "Movies",'fd': "Food",'tr': "Travel",'ee': "Electronics",'md': "Medical",'sp': "Shopping",'sv': "Services",'ot': "Others"}
	l3=[]
	dict21={}
	dates=[]
	#p=np.zeros((8,1))
	columns = [{} for i in range(8)]
	for record in query_set:
		if(record.borrower.username==name):
			dates.append(record.date)
	#print(len(dates))
	#np.resize(p,(8,len(dates)))
	#a=p.tolist()
	#print(dates) 
	for record in query_set:
		if(record.borrower.username==name):
			columns[dict2[record.tag]-1][record.date]=0.0

	for record in query_set:
		if(record.borrower.username==name):
			columns[dict2[record.tag]-1][record.date]=columns[dict2[record.tag]-1][record.date]+float(record.amount)

	#print(columns)
	x = [list(columns[i].keys()) for i in range(8)]
	y = [list(columns[i].values()) for i in range(8)]

	for record in query_set:
		if(record.borrower.username==name):
			dict21[record.tag]=dict2[record.tag]+float(record.amount)
	for (key, value) in dict21.items():
			dict21[key]=dict21[key]-dict2[key]
	newdict2={}
	for (key, value) in dict21.items():
		if value>0:
			newdict2[key] = value
	#print(newdict2)
	l2k=list(newdict2.keys())
	l2=[dictionary[i] for i in l2k]
	v2=list(newdict2.values())

	for record in query_set:
		if(record.borrower.username==name and record.lender.username!=name):
			l3.append(record.lender.username)
		elif(record.lender.username==name and record.borrower.username!=name):
			l3.append(record.borrower.username)
	dict3 = {k:0.0 for k in l3}
	# print("\n\ng")
	# print(l3)
	# print(dict3)
	for record in query_set:
		if(record.borrower.username==name and record.lender.username!=name):
			dict3[record.lender.username]=dict3[record.lender.username]+float(record.amount)
		elif(record.lender.username==name and record.borrower.username!=name):
			dict3[record.borrower.username]=dict3[record.borrower.username]+float(record.amount)

	fig = make_subplots(
	    rows=2, cols=2,
	    specs=[[{"type": "pie"}, {"type": "pie"}],
               [{"type": "bar"}, {"type": "bar"}]], 
	           subplot_titles=("Category-wise Expenditure", "Debts and Arrears with Friends", "Debts and Arrears")
	)
	# x = [datetime(year=2013, month=10, day=4),
	# datetime(year=2013, month=11, day=5),
	# datetime(year=2013, month=12, day=6)]

	# f = go.Figure(data=[go.Scatter(x=x, y=[1, 3, 6])])
	# # Use datetime objects to set xaxis range
	# f.update_layout(xaxis_range=[datetime(2013, 10, 17), datetime(2013, 11, 20)])

	# f2=go.FigureWidget(f)
	# fig.add_trace(f2)

	#labels = ["Movies", "Food", "Travel", "Electronics", "Medical","Shopping","Services","Others"]
	labels=l2
	values = v2#[v[2][1], v[2][2], v[2][3], v[2][4], v[2][5], v[2][6], v[2][7], v[2][8]]
	fig.add_trace(go.Pie(labels=labels,values=values),
	              row=1, col=1)

	labels = list(dict3.keys())
	values = list(dict3.values())
	fig.add_trace(go.Pie(labels=labels,values=values),
	              row=1, col=2)

	fig.add_trace(go.Bar(y=[2, 3, 1]),
	              row=2, col=1)

	fig.add_trace(go.Bar(y=[2, 3, 1]),
	              row=2, col=2)

	# fig.add_trace(go.Scatter3d(x=[2, 3, 1], y=[0, 0, 0], z=[0.5, 1, 2], mode="lines"),
	#               row=2, col=2)

	fig.update_layout(height=700, title_text="Insights")
	plot_div = plot(fig, output_type='div',include_plotlyjs=False, show_link=False, link_text="")

	fig1 = go.Figure()
	#for i in range(8):
	print("\n\n")
	print(x[1])
	print(y[1])
	fig1.add_trace(go.Scatter(x=x[0], y=y[0]))
	fig1.add_trace(go.Scatter(x=x[1], y=y[1]))
	fig1.add_trace(go.Scatter(x=x[2], y=y[2]))
	fig1.add_trace(go.Scatter(x=x[3], y=y[3]))
	fig1.add_trace(go.Scatter(x=x[4], y=y[4]))
	fig1.add_trace(go.Scatter(x=x[5], y=y[5]))
	fig1.add_trace(go.Scatter(x=x[6], y=y[6]))
	fig1.add_trace(go.Scatter(x=x[7], y=y[7]))
	#fig1.add_trace(go.Scatter(x=x[5], y=y[5]))
	#fig1 = go.Figure(data=[go.Scatter(x=x[1], y=y[1])])
	#fig = go.Figure(data=[go.Scatter(x=x, y=[1, 3, 6])])
# Use datetime objects to set xaxis range datetime.combine(date.today(), datetime.min.time())
	fig1.update_layout(xaxis_range=[datetime.combine(start_date, datetime.min.time()),datetime.combine(end_date, datetime.min.time())],title_text="Expenditure vs Time")
	plot_div2 = plot(fig1, output_type='div',include_plotlyjs=False, show_link=False, link_text="")
	animals=['giraffes', 'orangutans', 'monkeys']
	fig2 = go.Figure(data=[go.Bar(name='SF Zoo', x=animals, y=[20, 14, 23]),go.Bar(name='LA Zoo', x=animals, y=[12, 18, 29])
	])
# Change the bar mode
	fig2.update_layout(barmode='stack')
	plot_div3 = plot(fig2, output_type='div',include_plotlyjs=False, show_link=False, link_text="")

	#CHANGE

		# if(y.person2.username==f):
		# 	z = y.money_owed
		# 	a = str(y.person2)
	# template=loader.get_template('home.html')
	# x=''
	# context = {
	# 	'x' : x
	# }
	# #return HttpResponse(template.render(context, request))
	# return HttpResponse('success')
	return render(request, "insights.html", context={'plot_div1': plot_div, 'plot_div2':plot_div2, 'plot_div3':plot_div3})

